"""Lead persistence (Phase 13.3): a dedicated SQLite DB + Excel mirror.

Kept entirely separate from the courses DB and, deliberately, from anything that
touches OpenAI. This module is the only place captured PII is written.
"""

from __future__ import annotations

import os
import sqlite3
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

LEADS_DB = Path(os.getenv("LEADS_DB", "data/leads.sqlite"))
LEADS_XLSX = Path("data/leads.xlsx")
_XLSX_HEADERS = ["id", "name", "email", "phone", "course_slug", "session_id", "created_at"]


def _connect() -> sqlite3.Connection:
    LEADS_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(LEADS_DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_leads_db() -> None:
    with _connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS leads (
                id           TEXT PRIMARY KEY,
                form_id      TEXT UNIQUE NOT NULL,
                session_id   TEXT NOT NULL,
                name         TEXT NOT NULL,
                email        TEXT NOT NULL,
                phone        TEXT NOT NULL,
                course_slug  TEXT,
                source       TEXT NOT NULL DEFAULT 'chat',
                created_at   TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_leads_session ON leads(session_id);
            CREATE INDEX IF NOT EXISTS idx_leads_course  ON leads(course_slug);
            CREATE INDEX IF NOT EXISTS idx_leads_email   ON leads(email);
            """
        )
        conn.commit()


def get_lead_by_form_id(form_id: str) -> sqlite3.Row | None:
    init_leads_db()
    with _connect() as conn:
        return conn.execute("SELECT * FROM leads WHERE form_id = ?", (form_id,)).fetchone()


def insert_lead(
    *,
    form_id: str,
    session_id: str,
    name: str,
    email: str,
    phone: str,
    course_slug: str | None,
) -> tuple[str, bool]:
    """Insert a lead idempotently on form_id. Returns (lead_id, is_duplicate)."""
    init_leads_db()
    existing = get_lead_by_form_id(form_id)
    if existing:
        return existing["id"], True

    lead_id = str(uuid.uuid4())
    created_at = datetime.now(timezone.utc).isoformat()
    with _connect() as conn:
        conn.execute(
            """
            INSERT INTO leads (id, form_id, session_id, name, email, phone, course_slug, source, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'chat', ?)
            """,
            (lead_id, form_id, session_id, name, email, phone, course_slug, created_at),
        )
        conn.commit()

    try:
        append_to_leads_xlsx({
            "id": lead_id, "name": name, "email": email, "phone": phone,
            "course_slug": course_slug or "", "session_id": session_id, "created_at": created_at,
        })
    except Exception as e:  # excel mirror is best-effort; never fail the request
        import logging
        logging.getLogger(__name__).warning("leads excel mirror failed: %r", e)

    return lead_id, False


def _ensure_workbook() -> None:
    if LEADS_XLSX.exists():
        return
    LEADS_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "leads"
    ws.append(_XLSX_HEADERS)
    wb.save(LEADS_XLSX)


def append_to_leads_xlsx(row: dict) -> None:
    _ensure_workbook()
    wb = load_workbook(LEADS_XLSX)
    ws = wb["leads"]
    ws.append([row.get(h, "") for h in _XLSX_HEADERS])
    wb.save(LEADS_XLSX)
